"""
Loader per file Excel (.xlsx, .xls).

Excel è il formato più problematico per un sistema RAG perché i dati
sono TABULARI, non testuali. Un LLM non capisce una griglia di celle —
capisce frasi. Il lavoro del loader è trasformare le righe in frasi
leggibili che il modello possa comprendere.

Esempio di trasformazione:

    Foglio "Fornitori", riga 3:
    | Fornitore | Prodotto    | Prezzo/kg | Data       |
    | Ferrara   | Uva Italia  | 1.20      | 15/03/2025 |

    Diventa:
    "[Fornitori] Fornitore: Ferrara | Prodotto: Uva Italia |
     Prezzo/kg: 1.20 | Data: 15/03/2025"

Questa rappresentazione:
- Preserva il nome del foglio come contesto ([Fornitori])
- Associa ogni valore alla sua colonna (header)
- È leggibile sia dal modello che da un umano nei log
- Permette al retrieval di trovare "Ferrara" o "Uva Italia"

Ogni foglio diventa un Document separato. Un file con 3 fogli
produce 3 Document. Il motivo: fogli diversi contengono dati
diversi (fornitori vs vendite vs inventario) e il retrieval
deve poter attribuire la fonte correttamente.
"""

import openpyxl
from pathlib import Path

from rag_assistant.adapters.base import DocumentLoader
from rag_assistant.core.models import Document


class ExcelLoader(DocumentLoader):
    """Carica file Excel e converte ogni foglio in un Document."""

    def load(self, file_path: str) -> list[Document]:
        """Carica un file Excel.

        Ogni foglio (sheet) diventa un Document separato.
        La prima riga viene trattata come header e usata per
        dare contesto a ogni cella nelle righe successive.

        Returns:
            Lista di Document, uno per foglio non vuoto.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File non trovato: {file_path}")

        workbook = openpyxl.load_workbook(str(path), data_only=True)
        documents = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text = self._sheet_to_text(sheet, sheet_name)

            if not text.strip():
                continue

            documents.append(Document(
                # ID unico per foglio: path + nome foglio
                source_path=f"{path.resolve()}::{sheet_name}",
                source_name=f"{path.name} — {sheet_name}",
                doc_type="excel",
                text=text,
                metadata={
                    "file_name": path.name,
                    "sheet_name": sheet_name,
                    "total_rows": sheet.max_row or 0,
                    "total_cols": sheet.max_column or 0,
                },
            ))

        workbook.close()
        return documents

    def _sheet_to_text(self, sheet, sheet_name: str) -> str:
        """Converte un foglio Excel in testo leggibile.

        Strategia:
        1. La prima riga non vuota è l'header
        2. Ogni riga successiva diventa una frase:
           "[NomeFoglio] Header1: Valore1 | Header2: Valore2 | ..."
        3. Celle vuote vengono saltate
        4. Numeri, date e stringhe vengono tutti convertiti in stringa
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ""

        # Trova la prima riga non completamente vuota → è l'header
        headers = None
        data_start = 0

        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                headers = [str(cell).strip() if cell is not None else f"Col_{j}"
                           for j, cell in enumerate(row)]
                data_start = i + 1
                break

        if headers is None:
            return ""

        # Converti ogni riga in una frase leggibile
        lines = []
        for row in rows[data_start:]:
            # Salta righe completamente vuote
            if all(cell is None for cell in row):
                continue

            parts = []
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                header = headers[j] if j < len(headers) else f"Col_{j}"
                parts.append(f"{header}: {cell}")

            if parts:
                line = f"[{sheet_name}] {' | '.join(parts)}"
                lines.append(line)

        return "\n".join(lines)

    @staticmethod
    def supported_extensions() -> list[str]:
        return [".xlsx", ".xls"]
