"""Test per l'IngestionService.

Strategia: usiamo implementazioni fake di Embedder e VectorStore
(non Ollama e ChromaDB reali). Questo isola il test del service
dalla disponibilità dei servizi esterni.

Il pattern è lo stesso dei test del Blocco 1: creiamo sottoclassi
concrete delle ABC che si comportano in modo prevedibile.
"""

import pytest
import openpyxl

from rag_assistant.adapters.base import Embedder, VectorStore
from rag_assistant.core.models import Chunk, RetrievedChunk
from rag_assistant.services.ingestion_service import IngestionService


# ─── Fake Adapters ────────────────────────────────────────────────────────────

class FakeEmbedder(Embedder):
    """Embedder finto: restituisce vettori di lunghezza fissa.

    Tiene traccia di quante volte è stato chiamato, utile
    per verificare che il service lo usi correttamente.
    """

    def __init__(self, dimensions: int = 10):
        self.dimensions = dimensions
        self.call_count = 0
        self.texts_embedded: list[str] = []

    def embed(self, text: str) -> list[float]:
        self.call_count += 1
        self.texts_embedded.append(text)
        return [0.1] * self.dimensions

    def embed_batch(self, texts: list[str]) -> list[list[float]]:
        self.texts_embedded.extend(texts)
        self.call_count += len(texts)
        return [[0.1] * self.dimensions for _ in texts]


class FakeVectorStore(VectorStore):
    """Vector store in memoria per test.

    Salva tutto in una lista Python. Permette di verificare
    cosa è stato salvato senza toccare ChromaDB.
    """

    def __init__(self):
        self._chunks: list[Chunk] = []
        self._embeddings: list[list[float]] = []

    def add(self, chunks: list[Chunk], embeddings: list[list[float]]) -> None:
        if len(chunks) != len(embeddings):
            raise ValueError("Mismatch chunks/embeddings")
        self._chunks.extend(chunks)
        self._embeddings.extend(embeddings)

    def search(self, embedding: list[float], top_k: int = 5) -> list[RetrievedChunk]:
        results = []
        for c in self._chunks[:top_k]:
            results.append(RetrievedChunk(
                chunk_id=c.chunk_id,
                text=c.text,
                source_name=c.source_name,
                score=0.99,
            ))
        return results

    def clear(self) -> None:
        self._chunks = []
        self._embeddings = []

    def count(self) -> int:
        return len(self._chunks)

    @property
    def stored_chunks(self) -> list[Chunk]:
        """Accesso ai chunk salvati, per le assert nei test."""
        return self._chunks


# ─── Test ─────────────────────────────────────────────────────────────────────

class TestIngestionService:

    @pytest.fixture
    def embedder(self):
        return FakeEmbedder()

    @pytest.fixture
    def store(self):
        return FakeVectorStore()

    @pytest.fixture
    def service(self, embedder, store):
        return IngestionService(embedder=embedder, store=store)

    def test_ingest_txt_file(self, service, store, tmp_path):
        """Indicizza un file TXT e verifica che i chunk siano nello store."""
        # Crea un file con abbastanza testo da generare chunk
        text = " ".join(
            ["Questa è una frase di test per il chunking del documento."] * 50
        )
        file = tmp_path / "test.txt"
        file.write_text(text, encoding="utf-8")

        report = service.ingest_file(str(file))

        assert report["files_processed"] == 1
        assert report["files_failed"] == 0
        assert report["chunks_created"] > 0
        assert store.count() > 0

    def test_ingest_directory_mixed_formats(self, service, store, tmp_path):
        """Indicizza una directory con file di formati diversi."""
        # TXT
        txt_file = tmp_path / "note.txt"
        txt_file.write_text(
            " ".join(["Testo di esempio per il file di testo."] * 50),
            encoding="utf-8",
        )

        # Excel
        xlsx_file = tmp_path / "dati.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendite"
        ws.append(["Mese", "Totale"])
        for i in range(30):
            ws.append([f"Mese_{i}", 1000 + i * 100])
        wb.save(str(xlsx_file))

        # CSV
        csv_file = tmp_path / "report.csv"
        lines = ["Prodotto,Quantità,Prezzo"]
        for i in range(30):
            lines.append(f"Prodotto_{i},{100+i},{1.5+i*0.1:.1f}")
        csv_file.write_text("\n".join(lines), encoding="utf-8")

        report = service.ingest_directory(str(tmp_path))

        assert report["files_found"] == 3
        assert report["files_processed"] == 3
        assert report["files_failed"] == 0
        assert report["chunks_created"] > 0
        assert report["documents_created"] >= 3
        assert store.count() == report["chunks_created"]

    def test_ingest_directory_skips_unsupported(self, service, tmp_path):
        """I file con formato non supportato vengono ignorati."""
        (tmp_path / "immagine.png").write_bytes(b"fake png")
        (tmp_path / "note.txt").write_text(
            " ".join(["Testo valido per generare chunk."] * 50),
            encoding="utf-8",
        )

        report = service.ingest_directory(str(tmp_path))

        # Solo il TXT viene trovato e processato
        assert report["files_found"] == 1
        assert report["files_processed"] == 1

    def test_ingest_directory_not_found(self, service):
        """Directory inesistente solleva FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            service.ingest_directory("/non/esiste/")

    def test_ingest_directory_empty(self, service, tmp_path):
        """Directory vuota restituisce report con zero."""
        report = service.ingest_directory(str(tmp_path))

        assert report["files_found"] == 0
        assert report["files_processed"] == 0
        assert report["chunks_created"] == 0

    def test_ingest_file_not_found(self, service):
        """File inesistente solleva FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            service.ingest_file("/non/esiste.pdf")

    def test_corrupted_file_does_not_block_others(self, service, store, tmp_path):
        """Un file corrotto non blocca l'indicizzazione degli altri."""
        # File valido
        good_file = tmp_path / "buono.txt"
        good_file.write_text(
            " ".join(["Questo file è valido e deve essere indicizzato."] * 50),
            encoding="utf-8",
        )

        # File "corrotto" (estensione xlsx ma contenuto invalido)
        bad_file = tmp_path / "corrotto.xlsx"
        bad_file.write_bytes(b"questo non e' un excel valido")

        report = service.ingest_directory(str(tmp_path))

        assert report["files_found"] == 2
        assert report["files_processed"] == 1
        assert report["files_failed"] == 1
        assert len(report["errors"]) == 1
        assert "corrotto.xlsx" in report["errors"][0]["file"]
        assert store.count() > 0  # Il file buono è stato indicizzato

    def test_report_has_elapsed_time(self, service, tmp_path):
        """Il report contiene il tempo di esecuzione."""
        (tmp_path / "test.txt").write_text(
            " ".join(["Testo per il test dei tempi."] * 50),
            encoding="utf-8",
        )

        report = service.ingest_directory(str(tmp_path))

        assert "elapsed_seconds" in report
        assert report["elapsed_seconds"] >= 0

    def test_ingest_excel_multi_sheet(self, service, store, tmp_path):
        """Un Excel con più fogli genera più documenti."""
        xlsx_file = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Fornitori"
        ws1.append(["Nome", "Città"])
        for i in range(20):
            ws1.append([f"Fornitore_{i}", f"Città_{i}"])

        ws2 = wb.create_sheet("Clienti")
        ws2.append(["Nome", "Regione"])
        for i in range(20):
            ws2.append([f"Cliente_{i}", f"Regione_{i}"])

        wb.save(str(xlsx_file))

        report = service.ingest_file(str(xlsx_file))

        assert report["files_processed"] == 1
        assert report["documents_created"] == 2  # Due fogli = due documenti
        assert store.count() > 0

    def test_embedder_receives_chunk_texts(self, service, embedder, tmp_path):
        """L'embedder riceve esattamente i testi dei chunk."""
        file = tmp_path / "test.txt"
        file.write_text(
            " ".join(["Frase per verificare che l'embedder riceva il testo giusto."] * 50),
            encoding="utf-8",
        )

        service.ingest_file(str(file))

        # L'embedder deve aver ricevuto almeno un testo
        assert embedder.call_count > 0
        assert len(embedder.texts_embedded) > 0
        # Ogni testo embeddato deve essere non vuoto
        assert all(len(t) > 0 for t in embedder.texts_embedded)
